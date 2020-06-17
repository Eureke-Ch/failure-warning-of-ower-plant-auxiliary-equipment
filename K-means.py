import math
import matplotlib.pyplot as plt
import random
import xlrd
import xlwt
import openpyxl
def k_means(dataset, k, iteration):
    index = random.sample(list(range(len(dataset))), k)
    vectors = []
    for i in index:
        vectors.append(dataset[i])
    labels = []
    for i in range(len(dataset)):
        labels.append(-1)
    #根据迭代次数重复k-means聚类过程
    while(iteration > 0):
        C = []
        for i in range(k):
            C.append([])
        for labelIndex, item in enumerate(dataset):
            classIndex = -1
            minDist = 1e6
            for i, point in enumerate(vectors):
                dist = math.sqrt((item - point) ** 2)
                if(dist < minDist):
                    classIndex = i
                    minDist = dist
            C[classIndex].append(item)
            labels[labelIndex] = classIndex
        for i, cluster in enumerate(C):
            clusterHeart = 0
            for item in cluster:
                clusterHeart += item / len(cluster)
            vectors[i] = clusterHeart
        iteration -= 1
    return C, labels,vectors


def read_xsls(xlsx_path):
    data_xsls = xlrd.open_workbook(xlsx_path) 
    sheet_name = data_xsls.sheets()[0]  
    count_nrows = sheet_name.nrows  
    for i in range(1,count_nrows,20):
        # if (i<3000) or (i >390000 and i < 480000) or (i > 770000 and i < 810000):
        #     pass
        # else:
        for j in range(9):
            points[j].append(sheet_name.cell_value(i,j+1))

#读取初步处理的数据，并进行聚类
points = [[]for i in range(9)]
data_xsls = './data/14A17.xlsx'
read_xsls(data_xsls)
C = [];labels = [];vectors = []
for i in range(9):
    Ci,lablesi,vectorsi = k_means(points[i],5,50)
    for j in range(len(Ci)):
        if len(Ci[j]) == 0:
            Ci,lablesi,vectorsi = k_means(points[i],5,50)         
    C.append(Ci)
    labels.append(lablesi)
    vectors.append(vectorsi)


#将聚类完成的离散化数据保存下来

xls=openpyxl.Workbook()
sheet = xls.active
#sheet = xls.get_sheet_by_name('Sheet')
for i in range(len(labels[0])):
  for j in range(9):
    sheet.cell(i+2,j+2,labels[j][i]+(j+1)*100)
xls.save('li_san.xlsx')

path = 'vectors.xls'
rb = xlwt.Workbook()
sheet = rb.add_sheet('sheet1') 
for i in range(len(vectors[0])):
    for j in range(9):
        sheet.write(i,j,vectors[j][i])
rb.save(path)
'''
Colour = ['or','oy','og','ob','ok','ow','om']
for i in range(7):
    plt.plot(C[0][i],[0 for i in range(len(C[0][i]))],Colour[i])

plt.show()
'''