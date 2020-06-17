import xlrd
import xlwt
import openpyxl
import pylab as pl
from operator import itemgetter
from collections import OrderedDict,Counter
def read_xsls(xlsx_path):
    data_xsls = xlrd.open_workbook(xlsx_path) 
    sheet_name = data_xsls.sheets()[0]  
    count_nrows = sheet_name.nrows 
    for i in range(1,count_nrows,20):
      for j in range(9):
          data[j].append(sheet_name.cell_value(i,j+1))
      '''
      if i < 3000 or (i>390000 and i<480000) or (i > 770000 and i< 810000):
        pass
      else:
        for j in range(9):
          data[j].append(sheet_name.cell_value(i,j+1))
          '''
#读取数据
data_path1 = './data/14A17.xlsx'
data = [[]for i in range(9)]
read_xsls(data_path1) 

datasort = []
for i in range(9):
    datasort.append(data[i].copy())
for i in range(9):
    datasort[i].sort()
Q1 = [];Q3 = []
num = len(data[0])
for i in range(9):
    Q1.append(datasort[i][int(num/4)])
    Q3.append(datasort[i][int(num/4*3)])




xls=openpyxl.Workbook()
sheet = xls.active
#sheet = xls.get_sheet_by_name('Sheet')
pos = 0
for i in range(len(data[0])):
  isunusual = 0
  for k in range(9):
    if data[k][i] < Q1[k]-1.5*(Q3[k]-Q1[k]) or data[k][i] > Q3[k]+1.5*(Q3[k]-Q1[k]):
        isunusual = 1;pos += 1
        break
  if not isunusual:
    for j in range(9):
        sheet.cell(i+2-pos,j+2,data[j][i])

xls.save('./data/14ANew.xlsx')






























'''xls=openpyxl.Workbook()
sheet = xls.active
#sheet = xls.get_sheet_by_name('Sheet')
for i in range(len(data[0])):
  for j in range(9):
    sheet.cell(i+2,j+2,data[j][i])
xls.save('New.xlsx')'''