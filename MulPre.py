from numpy import *
import math
import xlrd
import numpy as np
import openpyxl
import datetime
def norm_pdf_multivariate(x, mu, sigma):
    size = len(x)
    if size == len(mu) and (size, size) == sigma.shape:
        det = linalg.det(sigma)
        if det == 0:
            raise NameError("The covariance matrix can't be singular")

        norm_const = 1.0/ (math.pow((2*pi),float(size)/2) * math.pow(det,1.0/2) )
        x_mu = matrix(x - mu)
        inv = sigma.I        
        result = math.pow(math.e, -0.5 * (x_mu * inv * x_mu.T))
        return norm_const * result
    else:
        raise NameError("The dimensions of the input don't match")

#读取清洗后的数据
datasigma = [];datatimeYes = []
dataabnormal = [];datatimeNo = []
data_xsls = xlrd.open_workbook("./data/14A17Process.xlsx") #打开此地址下的exl文档
#data_xsls = xlrd.open_workbook("13ANew.xls") 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows,5):
    if i < 3000 or (i>250000 and i<300000) or(i>380000 and i<490000) or (i > 760000 and i< 820000):
    #if i < 3000 or (i>50000 and i<100000) or (i > 330000 and i< 360000) or (i>850000 and i<990000):
        pass
    else:
        b = []
        for j in range(9):
            b.append(sheet_name.cell_value(i,j+1))
        datasigma.append(b)
        a = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
        datatimeYes.append(datetime.datetime(*a))
for i in range(1,count_nrows):
    if (i>397000 and i<400200) :
        b = []
        for j in range(9):
            b.append(sheet_name.cell_value(i,j+1))
        dataabnormal.append(b)
        a = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
        datatimeNo.append(datetime.datetime(*a))

#生成均值和协方差矩阵
sigma = matrix(cov(array(datasigma).T))
mean = []
for i in range(9):
    l = len(datasigma)
    mean.append(sum(datasigma[j][i]for j in range(l))/l)
mu = array(mean)


Y = []
for i in range(len(datasigma)):
    a = norm_pdf_multivariate(datasigma[i], mu, sigma)*pow(10,7)
    Y.append(a)
    

xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i] < 0.5:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,datasigma[i][j])
        sheet.cell(i+2-pos,1,datatimeYes[i])
    else:
        pos+=1
xls.save('./data/14A17normalN.xlsx')


Y = []
for i in range(len(dataabnormal)):
    a = norm_pdf_multivariate(dataabnormal[i], mu, sigma)*pow(10,7)
    Y.append(a)

xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i] < 0.5:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,dataabnormal[i][j])
        sheet.cell(i+2-pos,1,datatimeNo[i])
    else:
        pos+=1
xls.save('./data/14A17abnormalN.xlsx')
