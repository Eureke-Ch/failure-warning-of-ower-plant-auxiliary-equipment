import xlrd
import xlwt
import numpy as np
import openpyxl
import datetime
def read_xsls(xlsx_path):
    data_xsls = xlrd.open_workbook(xlsx_path) 
    sheet_name = data_xsls.sheets()[0]  
    count_nrows = sheet_name.nrows 
    for i in range(3,count_nrows):
        for j in range(7):
            if j == 0:
                a = xlrd.xldate_as_tuple(sheet_name.cell_value(i,j),0)
                data[j].append(datetime.datetime(*a))
            else:
                data[j].append(sheet_name.cell_value(i,j))
    for i in range(7):
        dataname.append(sheet_name.cell_value(0,i))
          
#读取数据
data = [[]for i in range(7)]
dataname = []
data_path1 = './data/14炉汽引风机18.xlsx'
read_xsls(data_path1) 
#data_path2 = './data/14炉汽引风机17-7-12.xlsx'
#read_xsls(data_path2)


mean = [[]for i in range(9)];std = [[]for i in range(9)]
for n in range(6):
    for i in range(0,len(data[0]),2880):
        if i+2880<len(data[0]):
            a = []
            for j in range(2880):
                a.append(data[n+1][i+j])
            meani = np.mean(a)
            mean[n].append(meani)
            # stdi = np.std(a)
            # std[n].append(stdi)
'''
stdsort = []
for i in range(9):
    stdsort.append(std[i].copy())
for i in range(9):
    stdsort[i].sort()
Q1 = [];Q3 = []
num = len(std[0])
for i in range(9):
    Q1.append(stdsort[i][int(num/4)])
    Q3.append(stdsort[i][int(num/4*3)])
stdnew = [[]for i in range(9)]
for i in range(9):
    for j in range(num):
        if std[i][j] > Q1[i]-1.5*(Q3[i]-Q1[i]) and std[i][j] < Q3[i]+1.5*(Q3[i]-Q1[i]):
            stdnew[i].append(std[i][j])

stdmean = []
for i in range(9):
    stdmean.append(np.mean(stdnew[i]))   
'''
for i in range(6):
    for j in range(2880,len(data[0])):
        day = int(j/2880)-1
        data[i+1][j] = (data[i+1][j]-mean[i][day])#/stdmean[i]
'''
datasort = []
for i in range(10):
    datasort.append(data[i].copy())
for i in range(9):
    datasort[i+1].sort()
Q1 = [];Q3 = []
num = len(data[0])
for i in range(9):
    Q1.append(datasort[i+1][int(num/4)])
    Q3.append(datasort[i+1][int(num/4*3)])
'''
xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(data[0])):
    '''
    isunusual = 0
    if (i<100000 and i>40000) or (i > 240000 and i< 270000) or (i>320000 and i<360000) or (i>840000 and i<990000):
        isunusual = 1;pos += 1
    else:
        for k in range(9):
            if data[k+1][i] < Q1[k]-1.5*(Q3[k]-Q1[k]) or data[k+1][i] > Q3[k]+1.5*(Q3[k]-Q1[k]):
                isunusual = 1;pos += 1
                break
    if not isunusual:
        for j in range(10):
            sheet.cell(i+2-pos,j+1,data[j][i])
    '''
    for j in range(7):
        sheet.cell(i+2-pos,j+1,data[j][i])
for i in range(7):
    sheet.cell(1,i+1,dataname[i])
xls.save('./data/14炉汽引风机18.xlsx')
print("处理完成")
