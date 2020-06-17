import numpy as np
import matplotlib.pyplot as plt
from sklearn.svm import LinearSVC
import xlrd
import datetime
import openpyxl
#%matplotlib inline
          
#读取数据
data_path1 = './data/14A17normal.xlsx'
data_path2 = './data/14A17abnormal.xlsx'
data = [];Y=[]
data_xsls = xlrd.open_workbook(data_path1) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows,10):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    Y.append(1)
    data.append(a)
data_xsls = xlrd.open_workbook(data_path2) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    for k in range(8):
        data.append(a)
        Y.append(0)



print("开始训练")

# solver='lbfgs',  MLP的求解方法：L-BFGS 在小数据上表现较好，Adam 较为鲁棒，SGD在参数调整较优时会有最佳表现（分类效果与迭代次数）；SGD标识随机梯度下降。
# alpha:L2的参数：MLP是可以支持正则化的，默认为L2，具体参数需要调整
# hidden_layer_sizes=(5, 2) hidden层2层,第一层5个神经元，第二层2个神经元)，2层隐藏层，也就有3层神经网络

clf = LinearSVC(random_state=0, tol=1e-5)
clf = clf.fit(data,Y)

Y = []
dataabnormal = [];datatime = []
data_xsls = xlrd.open_workbook("./data/14A18normal.xlsx") 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    dataabnormal.append(a)
    b = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
    datatime.append(datetime.datetime(*b))
    Y.append(clf.predict([a]))
    

xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i] == 0:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,dataabnormal[i][j])
        sheet.cell(i+2-pos,1,datatime[i])
    else:
        pos+=1
xls.save('./data/14A18PredictSVM.xlsx')