### load module
from sklearn import datasets
from sklearn.model_selection import train_test_split
from xgboost import XGBClassifier
from sklearn.metrics import accuracy_score
import xlrd
import datetime
import openpyxl
import numpy as np
### load datasets

data = [];target = []
data_xsls = xlrd.open_workbook("./data/14A17Process.xlsx") #打开此地址下的exl文档
#data_xsls = xlrd.open_workbook("13ANew.xls") 
sheet_name = data_xsls.sheets()[0]
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows,10):
    if i < 3000 or (i>390000 and i<490000) or (i > 770000 and i< 800000):
        pass
    else:
        b = []
        for j in range(9):
            b.append(sheet_name.cell_value(i,j+1))
        data.append(b)
        target.append(1)
for i in range(396603,403500):
    b = []
    for j in range(9):
        b.append(sheet_name.cell_value(i,j+1))
    data.append(b)
    target.append(0)


print("开始训练")

model = XGBClassifier()
model.fit(np.array(data),np.array(target))
Y = []
dataPre = [];datatime = []
data_xsls = xlrd.open_workbook("./data/14A18Process.xlsx") 
sheet_name = data_xsls.sheets()[0]  
#count_nrows = sheet_name.nrows  
for i in range(760000,800000):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    dataPre.append(a)
    b = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
    datatime.append(datetime.datetime(*b))
    Y.append(model.predict(np.array([a])))


xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i][0] != 10:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,dataPre[i][j])
        sheet.cell(i+2-pos,1,datatime[i])
        sheet.cell(i+2-pos,11,Y[i][0])
    else:
        pos+=1
xls.save('./data/14A18PredictXGboost.xlsx')
































'''

#读取数据
data_path1 = './data/14A17normal.xlsx'
data_path2 = './data/14A17abnormal.xlsx'
data = [];Y=[]
data_xsls = xlrd.open_workbook(data_path1) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows,10):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    Y.append(1)
    data.append(a)
data_xsls = xlrd.open_workbook(data_path2) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    for k in range(8):
        data.append(a)
        Y.append(0)

print("开始训练")

#digits = datasets.load_digits() 
### data analysis
#print(digits.data[0])
#print(digits.data.shape)   # 输入空间维度
#print(digits.target.shape) # 输出空间维度
### data split
#x_train,x_test,y_train,y_test = train_test_split(np.array(data),np.array(Y),test_size = 0.3,random_state = 33) 
### fit model for train data
model = XGBClassifier()
model.fit(np.array(data),np.array(Y))
Y = []
dataabnormal = [];datatime = []
data_xsls = xlrd.open_workbook("./data/14A18normal.xlsx") 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    dataabnormal.append(a)
    b = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
    datatime.append(datetime.datetime(*b))
    Y.append(model.predict(np.array([a])))
    

xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i] == 0:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,dataabnormal[i][j])
        sheet.cell(i+2-pos,1,datatime[i])
    else:
        pos+=1
xls.save('./data/14A18PredictXGboost.xlsx')

### make prediction for test data
y_pred = model.predict(x_test)
### model evaluate
accuracy = accuracy_score(y_test,y_pred)
print("accuarcy: %.2f%%" % (accuracy*100.0))
'''