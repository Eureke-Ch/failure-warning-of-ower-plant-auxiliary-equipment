#随机森林
from sklearn.ensemble import RandomForestClassifier
import xlrd
import matplotlib.pyplot as plt
import openpyxl
import datetime
#%matplotlib inline
          
#读取数据
data_path1 = './data/14A17normalRF.xlsx'
data_path2 = './data/14A17abnormalRF.xlsx'
data = [];Y=[]
data_xsls = xlrd.open_workbook(data_path1) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows,4):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    Y.append(1)
    data.append(a)
data_xsls = xlrd.open_workbook(data_path2) 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows 
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    data.append(a)
    Y.append(0)

clf = RandomForestClassifier().fit(data,Y)


Y = []
dataabnormal = [];datatime = []
data_xsls = xlrd.open_workbook("./data/14A18normalRF.xlsx") 
sheet_name = data_xsls.sheets()[0]  
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i,j+1))
    dataabnormal.append(a)
    b = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
    datatime.append(datetime.datetime(*b))
    Y.append(clf.predict([a]))
    

xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(Y)):
    if Y[i] == 0:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,dataabnormal[i][j])
        sheet.cell(i+2-pos,1,datatime[i])
    else:
        pos+=1
xls.save('./data/14A18PredictRF.xlsx')