import xlrd
import xlwt
import numpy as np
import openpyxl
def read_xsls(xlsx_path):
    data_xsls = xlrd.open_workbook(xlsx_path) 
    sheet_name = data_xsls.sheets()[0]  
    count_nrows = sheet_name.nrows 
    for i in range(1,count_nrows,10):
        for j in range(10):
            data[j].append(sheet_name.cell_value(i,j))
            datasort[j].append(sheet_name.cell_value(i,j))
    for i in range(10):
        dataname.append(sheet_name.cell_value(0,i))
          
#读取数据
data = [[]for i in range(10)]
datasort = [[]for i in range(10)]
dataname = []
data_path1 = './data/14A18-1-6.xlsx'
read_xsls(data_path1) 
data_path2 = './data/14A18-7-11.xlsx'
read_xsls(data_path2)

for i in range(9):
    datasort[i+1].sort()
Q1 = [];Q3 = []
num = len(data[0])
for i in range(9):
    Q1.append(datasort[i+1][int(num/4)])
    Q3.append(datasort[i+1][int(num/4*3)])


xls=openpyxl.Workbook()
sheet = xls.active
#sheet = xls.get_sheet_by_name('Sheet')
pos = 0
for i in range(len(data[0])):
    isunusual = 0
    #if (i<48000 and i>39000) or (i > 770000 and i< 810000):#14A17
    if i<100:
        isunusual = 1;pos += 1
    else:
        for k in range(9):
            if data[k+1][i] < Q1[k]-1.5*(Q3[k]-Q1[k]) or data[k+1][i] > Q3[k]+1.5*(Q3[k]-Q1[k]):
                isunusual = 1;pos += 1
                break
    if not isunusual:
        for j in range(10):
            sheet.cell(i+2-pos,j+1,data[j][i])
for i in range(10):
    sheet.cell(1,i+1,dataname[i])
xls.save('./data/14A18clean.xlsx')
print("清洗完成")




















'''
path = '13A17.xls'
rb = xlwt.Workbook()
sheet = rb.add_sheet('sheet1') 
for i in range(len(data[0])):
    for j in range(10):
        isunusual = 0
        for k in range(9):
            if data[k+1][i] < mean[k]-3*std[k] or data[k+1][i] > mean[k]+3*std[k]:
                isunusual = 1;break
        if not isunusual:
            sheet.write(i+1,j,data[j][i])

rb.save(path)
'''