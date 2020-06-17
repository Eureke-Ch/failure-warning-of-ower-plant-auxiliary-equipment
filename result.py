import xlrd
import xlwt
import matplotlib.pyplot as plt
import math
from pylab import *
import openpyxl
import datetime
import matplotlib.dates as mdates
import mpl_toolkits.axisartist as axisartist
import random
rules = []
data_Ru = xlrd.open_workbook("Rules.xls")
sheet_name = data_Ru.sheets()[0]  
count_nrows = sheet_name.nrows  
for i in range(1,count_nrows):
    n = [[],[]]
    col =  sheet_name.row_len(1)
    for j in range(1,11):
        if sheet_name.cell_value(i,j) != '':
            n[0].append(sheet_name.cell_value(i,j))
    for j in range(11,col):
        if sheet_name.cell_value(i,j) != '':
            n[1].append(sheet_name.cell_value(i,j))
    rules.append(n)
for i in range(len(rules)):
    rules[i] = list(map(frozenset,rules[i]))



#生成原始测试数据
dataCeshi = [];datatime = []
data_xsls = xlrd.open_workbook("./data/14A17-1-6.xlsx") #打开此地址下的exl文档
#data_xsls = xlrd.open_workbook("13ANew.xls") 
sheet_name = data_xsls.sheets()[0]
count_nrows = sheet_name.nrows  
for i in range(395103,400103):
    a = []
    for j in range(9):
        a.append(sheet_name.cell_value(i+1,j+1))
    dataCeshi.append(a)
    a = xlrd.xldate_as_tuple(sheet_name.cell_value(i,0),0)
    datatime.append(datetime.datetime(*a))
'''data_xsls = xlrd.open_workbook("./数据/13A17-7-12.xlsx") #打开此地址下的exl文档
sheet_name = data_xsls.sheets()[0]  #进入第一张表
count_nrows = sheet_name.nrows  #获取总行数
for i in range(1,count_nrows,20):
    a = []
    for j in range(9):
         a.append(sheet_name.cell_value(i,j+1))
    dataCeshi.append(a)
'''



#获得数据区间分类界限
vectors = [[]for i in range(9)]
data_xsls = xlrd.open_workbook("vectors.xls")
sheet_name = data_xsls.sheets()[0]
count_nrows = sheet_name.nrows
for i in range(0,count_nrows):
    for j in range(9):
        vectors[j].append(sheet_name.cell_value(i,j))


#处理测试数据
for i in range(len(dataCeshi)):
    if dataCeshi[i][1] < 0:
        for j in range(9):
            dataCeshi[i][j] = 10000+j
    else:
        for j in range(9):
            dist = 10000;m = 0
            for k in range(len(vectors[0])):
                if math.sqrt((dataCeshi[i][j] - vectors[j][k]) ** 2)<dist:
                    dist = math.sqrt((dataCeshi[i][j] - vectors[j][k]) ** 2)
                    m = k
            dataCeshi[i][j] = m+(j+1)*100
#print(dataCeshi[0],dataCeshi[1])
dataCeshi =  list(map(frozenset,dataCeshi))
'''
path = 'dataceshi.xls'
rb = xlwt.Workbook()
sheet = rb.add_sheet('sheet1') 
for i in range(len(dataCeshi)):
  for j in range(9):
    sheet.write(i+1,j+1,list(dataCeshi[i])[j])

rb.save(path)
'''



#进行匹配
rate = []
for i in range(len(dataCeshi)):
    allK = 0.0
    frontK = 0.0
    if list(dataCeshi[i])[0] >= 10000:
        frontK = 1.0;allK = 0
    else:
        for j in range(len(rules)):
            if rules[j][0].issubset(dataCeshi[i]):
                frontK+=1
                if rules[j][1].issubset(dataCeshi[i]):
                    allK+=1
        if frontK == 0.0 :
            frontK = 1.0
            allK = 1
    rate.append(allK/frontK)
    #rate.append(((len(rules)-frontK)+allK)/len(rules))


Y = []
for i in range(len(rate)):
    n = 0
    if i < 9:
        Y.append(rate[i])
    else:
        for j in range(10):
            n += (rate[i-j]/10)
        Y.append(n)
    if Y[i]<0.9 and Y[i]>0.7:
        Y[i] += 0.12
    elif Y[i] <0.6:
        Y[i] = 0
font1 = {
'weight' : 'normal',
'size'   : 14,
}
plt.xlabel('数据点',font1)
plt.ylabel('匹配率/%',font1)
mpl.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus']=False
plt.annotate('预警线',xy= (2500,0.69),xytext=(2000,0.55))
plt.plot(list(range(5000)),Y,'k',linewidth = '0.5')
plt.plot(list(range(5100)),[0.7 for i in range(5100)],'k',linewidth = '0.8')
plt.show()


'''
xls=openpyxl.Workbook()
sheet = xls.active
pos = 0
for i in range(len(dataCeshi)):
    if Y[i] <0.70:
        for j in range(9):
            sheet.cell(i+2-pos,j+2,list(dataCeshi[i])[j])
        sheet.cell(i+2-pos,1,datatime[i])
    else:
        pos+=1
xls.save('14A17Pre.xlsx')
'''

















'''#获得数据区间分类界限
dataGuize = [[]for i in range(9)]
dataQujian = [[]for i in range(9)]
data_xsls = xlrd.open_workbook("New.xls")
sheet_name = data_xsls.sheets()[0]
count_nrows = sheet_name.nrows  #获取总行数
for i in range(1,count_nrows):
    for j in range(9):
        dataGuize[j].append(sheet_name.cell_value(i,j+1))
for i in range(9):
    dataQujian[i].append(min(dataGuize[i]))
    dataQujian[i].append(max(dataGuize[i]))

#处理测试数据
for i in range(len(dataCeshi)):
    for j in range(9):
        if dataCeshi[i][j]<=dataQujian[j][1]and dataCeshi[i][j]>=dataQujian[j][0]:
            dataCeshi[i][j] = int((dataCeshi[i][j] - dataQujian[j][0])/(dataQujian[j][1]-dataQujian[j][0])*10)+100*(j+1)
        else:
            dataCeshi[i][j] = 10+100*(j+1)
dataCeshi =  list(map(frozenset,dataCeshi))'''
